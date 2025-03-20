import cv2
import customtkinter as ctk
from PIL import Image
import numpy as np
import threading
import os
import face_funcs
from datetime import datetime
from openpyxl import Workbook, load_workbook
from view_attendance import top_level

recorded_list = []
ENCODINGS = face_funcs.load_encodings('encoded_faces.dat')

# TODO: WHat happens with two faces????
# TODO: Test KOla's new model against his picture
# TODO: Attendance already taken for {label}


def process_webcam():
    global ver_img, ver_img_2
    
    # Open webcam
    ret, frame = video.read()

    if ret:
        # Convert BGR to RGBA
        opencv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        ver_img = opencv_image
        
        cv2.imshow('Webcam Output',ver_img)

        # Convert NumPy array to PIL Image
        captured_image = Image.fromarray(opencv_image)
        
        ver_img_2 = captured_image

        # Convert PIL Image to CTkImage for use in CTkLabel
        ctk_image = ctk.CTkImage(light_image=captured_image, size=(350, 350))

        # Update the label with the image
        label_widget.configure(image=ctk_image)
        label_widget.image = ctk_image  # Keep a reference to avoid garbage collection

    # Call this function again after 10 milliseconds
    label_widget.after(10, process_webcam)
    


def start_webcam_thread():
    threading.Thread(target=process_webcam, daemon=True).start()
    
def mark_attendance(name):
    
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%I:%M:%S %p")
    
    # Check if name has been recorded in runtime
    if name.lower in recorded_list:
        att_label.configure(text=f"Attendance already recorded for {name.title()}", text_color='#6A040F', font=ctk.CTkFont(family='Helvetica' ,size=12, weight='normal'))
        return
    # Checking if the file exists
    if os.path.exists('attendance.xlsx'):
        # Loads existing workbook
        wb = load_workbook('attendance.xlsx')
        ws = wb.active
    
    else:
        # Creates a new workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Date', 'Time'])

    ws.append([name, current_date, current_time])
    # TODO: Try to concatenate the name with the date/time sha
    wb.save('attendance.xlsx')

    recorded_list.append(name.lower)
    # TODO: Attendance Label
    att_label.configure(text=f"Attendance sucessfully recorded for {name.title()}", text_color="#3CB371")
    #print(f"Attendance sucessfully recorded for {name.title()}")
    
def verification():
    global ver_img_2
    
    # Convert ver_img_2 from PIL Image to NumPy array
    image = np.array(ver_img_2)
    
    # Ensure the image is in the right format (RGB)
    if image.shape[2] == 4:  # If the image has an alpha channel
        image = cv2.cvtColor(image, cv2.COLOR_RGBA2RGB)
    
    # Get face encodings
    try: 
        selected_label, face_tuple = face_funcs.verfiy_face(image, ENCODINGS)
    
    except TypeError: 
        # TypeError: cannot unpack non-iterable NoneType object
        # Usually occurs when no face is detected
        dec_txt.configure(text='No Face Detected', text_color='#FF4500')
        return  # Exit the function early since no valid face detection occurred
        
    # Reshape the captured image to match the model's input size
    resized_image = cv2.resize(image, (640, 640))
    
    # Update the detection in the UI
    dec_txt.configure(text=f'Detected Face: {selected_label}', text_color='white')
    dec_time.configure(text=f'Time: {datetime.now().strftime("%I:%M:%S %p")}')
    
    # Mark attendance
    mark_attendance(selected_label)
    
    if face_tuple:
        x1, y1, x2, y2 = face_tuple
        display_verified_face(x1, y1, x2, y2, selected_label)

            
def display_verified_face(x1,y1,x2,y2, label):
       
        # NOTE File testing image = cv2.imread('test.jpg')
        
        # XXX Using the frame from the camera feed
        global ver_img_2
        image = ver_img_2
        if isinstance(image, Image.Image):
            image = np.array(image)
            #image = 
        
        # Convert NumPy array to PIL Image
        # pil_image = Image.fromarray(image)
        
        
       # Draw the bounding box around the detected face/object    
        cv2.rectangle(image, (x1, y1), (x2, y2), (0, 255, 0), 2)
       
       # Display the label of the recognized object on the bounding box
        cv2.putText(image, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
        
        # Convert the OpenCV image (which is in BGR format) to RGB format
        #image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image_rgb = image

        # Convert the OpenCV image (Numpy Array) to a PIL image
        pil_image = Image.fromarray(image_rgb)
        
        ctk_verified_img = ctk.CTkImage(light_image=pil_image, size=(200, 300))
        ver_label = ctk.CTkLabel(root, text="", image=ctk_verified_img)
        
        ver_label.place(x=500, y=50)
        

# Create a CustomTkinter window
root = ctk.CTk()
root.geometry("850x500")
root.title("Attendance using Facial Recognition")
root.resizable(False, False) # Width, Height

title_label = ctk.CTkLabel(root, text='ATTENDANCE SYSTEM', font=ctk.CTkFont(size=30, weight='bold'))
title_label.pack()


# Feed Frame
feed_frame = ctk.CTkFrame(root, width=370, height=370, fg_color='#1034A6')

 
# Result Frame
result_frame = ctk.CTkFrame(root, )
 
 
# Live feed label
label_widget = ctk.CTkLabel(root, text="")
label_widget.place(x=70, y=50)

# Initialize the webcam outside of the function
video = cv2.VideoCapture(0)
#address = "http://10.157.104.106:8080/video"
#video.open(address)
feed_frame.place(x=60,y=40)

# Start the webcam processing in a separate thread
start_webcam_thread()


# ResNet -- deep learning model for extracting facial landmarks

# Verify button
verify_button = ctk.CTkButton(root, text='VERIFY FACE', width=350, command=verification)
verify_button.place(x=70, y=420)

# Detection status
dec_txt = ctk.CTkLabel(root, text='', font=ctk.CTkFont(family='Helvetica' ,size=12, weight='normal'))
dec_txt.place(x=500, y=420)

dec_time = ctk.CTkLabel(root, text='', font=ctk.CTkFont(family='Helvetica' ,size=12, weight='normal'))
dec_time.place(x=500, y=440)

# Successful Attendance
att_label = ctk.CTkLabel(root, text='', font=ctk.CTkFont(family='Helvetica' ,size=12, weight='normal'), text_color='#3CB371')
att_label.place(x=500, y=460)

att_button = ctk.CTkButton(root, text="VIEW ATTENDANCE", width=350, command=top_level)
att_button.place(x=70, y=450)

# Run the Tkinter main loop
root.mainloop()

# Release the webcam on exit
video.release()