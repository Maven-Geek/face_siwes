import cv2
import face_recognition
import numpy as np
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


def save_encodings(encoded_dict):
    with open('encoded_faces.dat', 'wb') as f:
        pickle.dump(encoded_dict, f)
        return f'Encodings Saved'

def load_encodings(file_name):
    with open(file_name, 'rb') as f:
        return pickle.load(f)

def create_encodings_dict(path):
    '''Returns a dict of encodings of the images in a folder path with the name of the image as the key of the dict'''

    enc_dict = {}
    img_list = os.listdir(path)
    for i in img_list:
        #print(i.split('.')[0])
        name = i.split('.')[0]
        img_full_path = os.path.join(path,i)
        the_image = cv2.imread(img_full_path)
        img_face_loc = face_recognition.face_locations(the_image, model='cnn')
        enc_dict[name] = face_recognition.face_encodings(the_image, img_face_loc)
    return enc_dict

#save_encodings(create_encodings_dict("load_images"))

def draw_box(frame, frame_loc):
    '''Draws a box on the face in the image'''
    # Iterate over each face detected
    for (top, right, bottom, left) in frame_loc:
        # Define the start and end points for the rectangle
        start_point = (left, top)  # (x, y) = (left, top)
        end_point = (right, bottom)  # (x, y) = (right, bottom)

        # Blue color in BGR
        color = (255, 0, 0)
        

        # Line thickness of 2 px
        thickness = 2

        # Draw the rectangle on the frame
        cv2.rectangle(frame, start_point, end_point, color, thickness)

def compare_images(encodings_file, stream_input):
    enc_dict = encodings_file #load_encoding(encodings_file)
    stream_enc = face_recognition.face_encodings(stream_input)

    #print(stream_enc[0])
    #print()
    
    for name, enc in enc_dict.items():
        #enc_list = list(end_dict.values())
        #print(enc)
        if len(stream_enc) > 0:
            result = face_recognition.compare_faces(enc, stream_enc[0], tolerance=0.6)[0]
            distance = face_recognition.face_distance(enc, stream_enc[0])
        #print(distance)  --- Can be used for the best match index
        # along with np.argmin(distances)
       # print(result)

            # print facial distance
            if result == True:
                mark_attendance(name)
        else:
            print("No face found in this frame")
        

def mark_attendance(name):
    
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%I:%M:%S %p")
    recorded_list = []

    # Check is name exists  current runtime
    if name in recorded_list:
        pass
    # Checking if the file exists
    if os.path.exists('attendance.xlsx'):
        # Loads existing workbook
        wb = load_workbook('attendance.xlsx')
        ws = wb.active
    
    else:
        # Creates a new workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Date', 'Time'])

    ws.append([name, current_date, current_time])
    # TODO: Try to concatenate the name with the date/time sha
    wb.save('attendance.xlsx')

    
    print(f"Attendance sucessfully recorded for {name.title()}")

def verfiy_face(imgS, encodings):
# Detect face locations and encodings in the current frame
        att_enc = encodings
        known_enc = list(att_enc.values())
        known_enc_list = [i[0] for i in known_enc]
        # XXX imgS = cv2.resize(imgS, (0, 0), fx=0.5, fy=0.5)
        faceCurFrame = face_recognition.face_locations(imgS, model='hog')
        # faceCurFrame = imgS
        encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        # Iterate over each detected face
        for enc_face, face_loc in zip(encodeCurFrame, faceCurFrame):
            # Compare the current face with known faces
            matches = face_recognition.compare_faces(known_enc_list, enc_face,) # FIXME Unknown face detection: set the tolerance
            face_dist = face_recognition.face_distance(known_enc_list, enc_face)
            print("Matches:", matches)
            print("Distances:", face_dist)
            
            # TODO -- Unknown face detected
            # I intend to use if matches = False (i.e if it's empty)
            
            match_index = np.argmin(face_dist)
            print(f"Match Index: {list(att_enc.keys())[match_index]}")
            label = list(att_enc.keys())[match_index]
            y1, x2, y2, x1 = face_loc
            # Multiplied by 4 to upscale the box
            upscale_factor = int(1 / 1)
            return label,(x1*upscale_factor,y1*upscale_factor,x2*upscale_factor,y2*upscale_factor) 

            
            # if matches[match_index]:
            #     # If a match is found, draw a rectangle around the face
            #     y1, x2, y2, x1 = face_loc
            #     y1, x2, y2, x1 = y1 * 2, x2 * 2, y2 * 2, x1 * 2  # Scale back to original size
                
            #     # Draw rectangle around the face
            #     start_point = (x1, y1)
            #     end_point = (x2, y2)
            #     color = (255, 0, 0)  # Blue color in BGR
            #     thickness = 2  # Thickness of the rectangle
            #     cv2.rectangle(frame, start_point, end_point, color, thickness)